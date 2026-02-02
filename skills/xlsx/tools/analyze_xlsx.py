#!/usr/bin/env python3
"""
Analyze Excel workbook structure.

Usage:
    python analyze_xlsx.py --file data.xlsx
"""

import argparse
import sys
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({
        "status": "error",
        "error": "'openpyxl' package is required. Install with: pip install openpyxl",
        "message": "Missing dependency"
    }))
    sys.exit(1)


def analyze_xlsx(file_path: str) -> dict:
    """Analyze workbook structure."""
    wb = load_workbook(file_path)
    
    sheets_info = []
    total_cells = 0
    total_formulas = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        cell_count = 0
        formula_count = 0
        formulas = []
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_count += 1
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                        if len(formulas) < 5:  # Limit examples
                            formulas.append({
                                "cell": cell.coordinate,
                                "formula": cell.value
                            })
        
        sheets_info.append({
            "name": sheet_name,
            "dimensions": f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
            "rows": ws.max_row,
            "columns": ws.max_column,
            "cell_count": cell_count,
            "formula_count": formula_count,
            "sample_formulas": formulas
        })
        
        total_cells += cell_count
        total_formulas += formula_count
    
    return {
        "file": str(file_path),
        "sheet_count": len(wb.sheetnames),
        "total_cells": total_cells,
        "total_formulas": total_formulas,
        "sheets": sheets_info
    }


def main():
    parser = argparse.ArgumentParser(description='Analyze Excel workbooks')
    parser.add_argument('--file', required=True, help='Path to Excel file')
    
    args = parser.parse_args()
    
    if not Path(args.file).exists():
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)
    
    try:
        result = analyze_xlsx(args.file)
        
        output = {
            "status": "success",
            "data": result,
            "message": "Successfully analyzed workbook"
        }
        
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        print(json.dumps(output, indent=2))
        
    except Exception as e:
        print_json_error(str(e))

def print_json_error(message, exit_code=1):
    result = {
        "status": "error",
        "error": message,
        "message": message
    }
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    print(json.dumps(result, ensure_ascii=False))
    sys.exit(exit_code)


if __name__ == '__main__':
    main()
