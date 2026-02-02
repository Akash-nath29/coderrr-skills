#!/usr/bin/env python3
"""
Create Excel workbooks.

Usage:
    python create_xlsx.py --output data.xlsx --sheets '[...]'
"""

import argparse
import sys
import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({
        "status": "error",
        "error": "'openpyxl' package is required. Install with: pip install openpyxl",
        "message": "Missing dependency"
    }))
    sys.exit(1)


def create_xlsx(output_path: str, sheets_spec: list):
    """Create an Excel workbook."""
    wb = Workbook()
    
    # Remove default sheet if we're creating new ones
    if sheets_spec:
        default_sheet = wb.active
        wb.remove(default_sheet)
    
    for sheet_spec in sheets_spec:
        sheet_name = sheet_spec.get('name', 'Sheet1')
        ws = wb.create_sheet(title=sheet_name)
        
        row_num = 1
        
        # Add headers
        headers = sheet_spec.get('headers', [])
        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
            row_num += 1
        
        # Add data
        data = sheet_spec.get('data', [])
        for row_data in data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                if isinstance(value, str) and value.startswith('='):
                    cell.value = value  # Formula
                else:
                    cell.value = value
            row_num += 1
        
        # Set column widths
        col_widths = sheet_spec.get('column_widths', {})
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Create Excel workbooks')
    parser.add_argument('--output', required=True, help='Output file path (.xlsx)')
    parser.add_argument('--sheets', required=True, help='JSON specification of sheets')
    
    args = parser.parse_args()
    
    try:
        sheets = json.loads(args.sheets)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid sheets JSON - {e}", file=sys.stderr)
        sys.exit(1)
    
    try:
        result = create_xlsx(args.output, sheets)
        
        output = {
            "status": "success",
            "data": {
                "file": result,
                "sheets": len(sheets)
            },
            "message": "Successfully created workbook"
        }
        
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        print(json.dumps(output, indent=2))
        
    except Exception as e:
        print_json_error(str(e))

def print_json_error(message, exit_code=1):
    result = {
        "status": "error",
        "error": message,
        "message": message
    }
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    print(json.dumps(result, ensure_ascii=False))
    sys.exit(exit_code)


if __name__ == '__main__':
    main()
